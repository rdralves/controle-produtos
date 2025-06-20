from openpyxl.utils import get_column_letter
import openpyxl
from django.http import HttpResponse
from django.views.generic import ListView
from django.views import View
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Produto


@login_required
def produto_list(request):
    produtos = Produto.objects.all()
    return render(request, 'estoque/produto_list.html', {'produtos': produtos})


class ProdutoListView(LoginRequiredMixin, ListView):
    model = Produto
    template_name = 'estoque/produto_list.html'
    context_object_name = 'produtos'


class ProdutoDetailView(LoginRequiredMixin, DetailView):
    model = Produto
    template_name = 'estoque/produto_detail.html'


class ProdutoCreateView(LoginRequiredMixin, CreateView):
    model = Produto
    fields = ['ref_produto', 'descricao', 'quantidade', 'valor']
    template_name = 'estoque/produto_form.html'
    success_url = reverse_lazy('produto_list')


class ProdutoUpdateView(LoginRequiredMixin, UpdateView):
    model = Produto
    fields = ['ref_produto', 'descricao', 'quantidade', 'valor']
    template_name = 'estoque/produto_form.html'
    success_url = reverse_lazy('produto_list')


class ProdutoDeleteView(LoginRequiredMixin, DeleteView):
    model = Produto
    template_name = 'estoque/produto_confirm_delete.html'
    success_url = reverse_lazy('produto_list')


# views.py


class RelatorioProdutosView(LoginRequiredMixin, ListView):
    model = Produto
    template_name = 'estoque/relatorio_produtos.html'
    context_object_name = 'produtos'


class ExportarProdutosExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        produtos = Produto.objects.all()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Produtos'

        headers = ['Referência', 'Descrição',
                   'Quantidade', 'Valor', 'Data Entrada']
        worksheet.append(headers)

        for produto in produtos:
            worksheet.append([
                produto.ref_produto,
                produto.descricao,
                produto.quantidade,
                float(produto.valor),
                produto.data_entrada.strftime('%d/%m/%Y')
            ])

        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=produtos.xlsx'
        workbook.save(response)
        return response
